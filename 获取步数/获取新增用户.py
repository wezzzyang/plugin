import json
from openpyxl import load_workbook
import os


script_directory = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "userGameProcess.xlsx",
)

df = load_workbook(script_directory)


print(df[df.sheetnames[0]])

amount_obj = {}


for row in df[df.sheetnames[0]].iter_rows(
    values_only=True
):  # values_only=True 返回单元格的值（而非单元格对象）
    if row[2] == "sync_data":
        continue

    i = json.loads(row[2])["user_step"]

    if not i:
        continue

    if amount_obj.get(i) == None:
        amount_obj[i] = 1
    else:
        amount_obj[i] += 1

print(amount_obj)
